"""
Support for openpyxl

Use of module functions is not elegant.  I should try to subclass
openpyxl.workbook.Workbook

This works for open-pyxl 1.x.  Starting with 2.0, the row and column indexing
starts with 1 instead of 0.
"""
import logging
from openpyxl import workbook
from openpyxl import load_workbook
import sys
from openpyxl.cell import get_column_letter
#from openpyxl.worksheet import cells_from_range

module_logger = logging.getLogger(__name__)

def getValueWithMergeLookup(sheet, col, row):
  """
  from http://stackoverflow.com/questions/23562366/\
                                    how-do-i-get-value-present-in-a-merged-cell
  does not work in this version of openpyxl
  """
  idx = '{0}{1}'.format(get_column_letter(col), row)
  for range_ in sheet.merged_cell_ranges:
        cells = list(cells_from_range(range_))[0]
        if idx in cells:
            # If this is a merged cell, you can look up the value
            # in the first cell of the merge range
            return sheet.cell(cells[0]).value
  return sheet.cell(row=row, column=col).value

def get_column_names(sh):
  """
  Dictionary of column names indexed by column letter

  @param sh : worksheet
  @type  sh : Worksheet object
  """
  size = sh.calculate_dimension()
  maxrow = sh.get_highest_row()
  maxcol = sh.get_highest_column()
  column_names = {}
  for col_num in xrange(maxcol):
    #col = get_column_letter(col_num_LSTs)
    column_names[col_num] = sh.cell(column=col_num,row=0).value
  return column_names

def get_column_id(sh,col_name):
  """
  Returns column letter of the first column with that name

  @param sh : worksheet

  @param col_name : column name

  @return: int
  """
  sys.stdout.flush()
  col_names = get_column_names(sh)
  sys.stdout.flush()
  for col_id in col_names.keys():
    sys.stdout.flush()
    if col_names[col_id] == col_name:
      return col_id
  # not found
  #module_logger.error("get_column_id: %s not found", col_name)
  return None

def find_column(sh,name):
  """
  Find or create a column
  """
  col = get_column_id(sh,name)
  if col == None:
    #module_logger.error("Column %s not found",name,exc_info=True)
    col = sh.get_highest_column()
    sh.cell(column= col, row=0).value = name
  return col

def get_column(ws, column_name):
  """
  Get the data as a list from a column identified by its name.

  @param ws : worksheet

  @param column_name : column name
  @type  column_name : str

  @return: list of data
  """
  column_id = get_column_id(ws,column_name)
  module_logger.debug("get_column: column ID for %s is %s",
                      column_name, column_id)
  if column_id != None:
    column = ws.columns[column_id]
    #module_logger.debug("get_column: column label is %s", column[0])
    column_data = []
    # The first cell always has the label
    for cell in column[1:]:
      column_data.append(cell.value)
    return column_data
  else:
    #module_logger.debug("get_column: failed")
    return None

def insert_empty_row_after(ws,prior_row):
  """
  Move all the rows above 'prior_row' down one and 'prior_row'+1 emptied.

  @param ws : worksheet

  @param prior_row : the row below the one to be emptied
  @type  prior_row : a tuple of cells

  @return: True
  """
  highest_row = ws.get_highest_row()
  highest_col = ws.get_highest_column()
  max_col_letter = get_column_letter(highest_col)
  print max_col_letter
  for row in range(highest_row,prior_row,-1):
    module_logger.debug("insert_empty_row_after: processing row %d",row)
    for col in range(highest_col):
      ws.cell(row=row+1,column=col).value = ws.cell(row=row,column=col).value
  for col in range(highest_col):
    ws.cell(row=prior_row+1,column=col).value = None
  return True

def get_row_number(ws,column,value):
  """
  Get the first row which has the specified value in the designated column

  @param ws : worksheet
  @type  ws : Worksheet() object

  @param column : column ID
  @type  column : int

  @param value : value to be matched by contents
  """
  highest_row = ws.get_highest_row()
  for row in range(highest_row):
    cell_id = "%s%s" % (chr(column+65),str(row+1))
    if ws.cell(cell_id).value == value:
      return row
      break
  return None

def column_id(index):
  """
  Return the column label given its numerical index

  @param index : column number, with the left column being column 0
  @type  index : int

  @return: one or two character letter code
  """
  if index < 26:
    col_letter = chr(65+index)
  else:
    col_letter = chr(64+(index/26))+chr(65+(index%26))
  module_logger.debug("column_id: column %d has code %s",index,col_letter)
  return col_letter

def column_number(code):
  """
  Return column number, with left column being 0.

  @param code : column letter code
  @type  code : str

  @return int
  """
  if len(code) == 1:
    return ord(code)-65
  elif len(code) == 2:
    return 26*(ord(code[0])-64) + ord(code[1])-65

def set_column_dimensions(ws):
  """
  Set the width of the columns to fit the data in the top two rows

  @param ws : worksheet
  """
  for index in range(ws.get_highest_column()):
    width = max(len(str(ws.cell(row=0,column=index).value)),
                len(str(ws.cell(row=1,column=index).value)))
    module_logger.debug("set_column_dimensions: processing column %d",index)
    sys.stdout.flush()
    col_letter = column_id(index)
    ws.column_dimensions[col_letter].width = width+1
  return True

def column_ID_dict(worksheet):
  """
  Create a dictionary of column number indexed by column name
  """
  column_names = get_column_names(worksheet)
  column_index = {}
  module_logger.debug("column_ID_dict: %s has columns: %s",
                 worksheet.title, str(column_names))
  for col in column_names.keys():
    column_index[column_names[col]]  = get_column_id(worksheet,
                                                     column_names[col])
  return column_names, column_index

def get_rows(sh,colname,value):
  """
  Get a list of rows (each a list of cells) with a given value in a column
  """
  rows = []
  col_id = get_column_id(sh,colname)
  for row in sh.rows:
    if row[col_id].value == value:
      rows.append(row)
  return rows

def get_row_numbers(sh,colname,value):
  """
  Get the row  numbers for rows with a given value in a column

  This is like 'get_rows' but returns row numbers, not a list of rows.
  """
  rows = []
  col_id = get_column_id(sh,colname)
  for row in range(1,sh.get_highest_row()):
    if sh.cell(row=row,column=col_id).value == value:
      rows.append(row)
  return rows

def add_column(sheet, meta_column, col, colname):
  """
  Add one empty column to a sheet

  @param sheet : sheet to be updated

  @param col : column number

  @param colname : text for row 0 in the column
  """
  sheet.cell(column=col,row=0).value = colname
  meta_column[colname] = col
  return meta_column

def add_columns(sheet,names):
  """
  Add a list of empty columns to a sheet

  @param sheet : sheet to be updated

  @param names : names of the new columns
  @type  names : list of str
  """
  # Does the column exist?
  num_cols = len(names)
  for index in range(num_cols):
    col_id = get_column_id(sheet,names[index])
    if col_id == None:
      col = sheet.get_highest_column()
    else:
      col = column_number(col_id)
    bf_logger.debug("add_columns: %s is column %d", names[index], col)
    sys.stdout.flush()
    add_column(sheet,col,names[index])


class Workbook(workbook.Workbook):
  """
  incomplete
  """
  def __init__(self, path="./", file=None):
    """
    Create an instance of FirmwareServer()
    """
    if file:
      self = load_workbook(path+file)
    self.path = path
    self.file = file
    self.logger = logging.getLogger(module_logger.name+".Workbook")
